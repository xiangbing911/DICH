import argparse
import os
import re
import sys
from datetime import datetime
from typing import Optional, Tuple


def _safe_name(text: str) -> str:
    return re.sub(r"[^0-9A-Za-z\u4e00-\u9fff]+", "_", text).strip("_")


def _resolve_film_type_args(
    *, film_type_cell: str
) -> Tuple[Optional[str], Optional[str]]:
    """
    Returns (film_type_key, film_type_custom).
    - If matches a preset key or preset label, use --film-type.
    - Otherwise use --film-type-custom with original cell.
    """
    import qwen_feiyi_to_csv as gen

    raw = (film_type_cell or "").strip()
    if not raw:
        return "documentary", None

    if raw in gen.FILM_TYPE_PRESETS:
        return raw, None

    for k, label in gen.FILM_TYPE_PRESETS.items():
        if raw == label:
            return k, None

    return None, raw


def main(argv: Optional[list[str]] = None) -> int:
    p = argparse.ArgumentParser(
        description="Read 非遗项目汇总.xlsx (项目名称/短片类型/背景资料) and generate one CSV per row."
    )
    p.add_argument("--xlsx", default="非遗项目汇总.xlsx", help="输入 XLSX 路径")
    p.add_argument("--sheet", default=None, help="Sheet 名称（默认第一个）")
    p.add_argument(
        "--start-row",
        type=int,
        default=2,
        help="数据起始行（默认 2，跳过表头）",
    )
    p.add_argument(
        "--out-dir",
        default=os.path.join("outputs", "batch"),
        help="输出目录（默认 outputs/batch）",
    )
    p.add_argument("--model", default="qwen3-max", help="Qwen model name")
    p.add_argument(
        "--base-url",
        default="https://dashscope.aliyuncs.com/compatible-mode/v1",
        help="DashScope OpenAI-compatible base URL",
    )
    p.add_argument("--api-key", default=None, help="API Key（可选，默认读 DASHSCOPE_API_KEY）")
    p.add_argument("--no-stream", action="store_true", help="Disable streaming output")
    p.add_argument(
        "--scene-count",
        type=int,
        default=None,
        help="场景数量（可选；开启自动规划时仅作为兜底）",
    )
    p.add_argument(
        "--scene-count-col",
        type=int,
        default=4,
        help="场景数量所在列（默认 4）",
    )
    p.add_argument(
        "--auto-plan",
        action="store_true",
        default=True,
        help="由模型自动决定短片类型与场景数量（默认开启）",
    )
    p.add_argument(
        "--no-auto-plan",
        action="store_false",
        dest="auto_plan",
        help="关闭自动规划，按表格/参数指定短片类型与场景数",
    )
    args = p.parse_args(argv)

    try:
        import openpyxl  # type: ignore
    except Exception as e:
        print(
            "Missing dependency: openpyxl. Install via: pip install openpyxl\n"
            f"Import error: {e}",
            file=sys.stderr,
        )
        return 2

    import qwen_feiyi_to_csv as gen

    if not os.path.exists(args.xlsx):
        print(f"XLSX not found: {args.xlsx}", file=sys.stderr)
        return 2

    wb = openpyxl.load_workbook(args.xlsx, read_only=True, data_only=True)
    ws = wb[args.sheet] if args.sheet else wb.worksheets[0]

    os.makedirs(args.out_dir, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    processed = 0
    file_seq = 0
    failed = 0
    for row_idx in range(args.start_row, ws.max_row + 1):
        name = ws.cell(row=row_idx, column=1).value
        film_type_cell = ws.cell(row=row_idx, column=2).value
        background = ws.cell(row=row_idx, column=3).value
        scene_count_cell = (
            ws.cell(row=row_idx, column=args.scene_count_col).value
            if args.scene_count_col >= 1
            else None
        )

        topic = str(name or "").strip()
        if not topic:
            continue

        film_type_key, film_type_custom = _resolve_film_type_args(
            film_type_cell=str(film_type_cell or "")
        )
        if args.auto_plan:
            film_type_key, film_type_custom = None, None
        background_text = str(background or "").strip() or None

        file_seq += 1
        safe_topic = _safe_name(topic)
        out_path = os.path.join(args.out_dir, f"{file_seq:04d}_{safe_topic}_{ts}.csv")

        row_scene_count = None
        if scene_count_cell is not None:
            try:
                row_scene_count = int(str(scene_count_cell).strip())
            except Exception:
                row_scene_count = None

        print(f"\n=== Row {row_idx}: {topic} / {film_type_custom or film_type_key or ''} ===")
        try:
            film_type: Optional[str] = None
            film_preset_key: Optional[str] = None
            if not args.auto_plan:
                film_type, film_preset_key = gen.resolve_film_type(
                    film_type_preset=film_type_key, film_type_custom=film_type_custom
                )
            if row_scene_count and row_scene_count > 0:
                scene_count = int(row_scene_count)
            elif args.scene_count is not None:
                scene_count = int(args.scene_count)
            else:
                scene_count = None
            out_csv, story_title, _story_content, _warnings, decision = gen.generate_csv(
                topic=topic,
                model=args.model,
                base_url=args.base_url,
                api_key=args.api_key,
                stream=(not args.no_stream),
                film_type=film_type,
                film_preset_key=film_preset_key,
                scene_count=scene_count,
                auto_plan=bool(args.auto_plan),
                background=background_text,
                out_path=out_path,
            )
            processed += 1
            print(f"Saved CSV: {out_csv}")
            print(f"Story title: {story_title}")
            print(f"Film type: {decision.film_type}")
            print(f"Scene count: {decision.scene_count}")
        except Exception as e:
            failed += 1
            print(f"[ERROR] Row {row_idx} failed: {e}", file=sys.stderr)

    print(f"Done. ok={processed} failed={failed} out_dir={args.out_dir}")
    return 0 if failed == 0 else 1


if __name__ == "__main__":
    raise SystemExit(main())
